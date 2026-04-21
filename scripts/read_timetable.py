import openpyxl
import json

wb = openpyxl.load_workbook('d:/VSCode_Workspace/jetlag_snake/ktx_timetable.xlsx')

# Just get sheet names and dimensions
summary = {}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    summary[sheet] = {"rows": ws.max_row, "cols": ws.max_column}

with open('d:/VSCode_Workspace/jetlag_snake/scripts/sheet_summary.json', 'w', encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)

# For each sheet, get the English station name row and first 3 data rows
details = {}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Find row with English station names (contains "Seoul" or "Yongsan" or similar)
    header_row_idx = None
    english_row_idx = None
    for i, row in enumerate(all_rows):
        row_str = str(row)
        if 'Seoul' in row_str or 'Yongsan' in row_str or 'Daejeon' in row_str:
            english_row_idx = i
        if '열차번호' in row_str:
            header_row_idx = i
            break

    if header_row_idx is not None:
        details[sheet] = {
            "header_row": header_row_idx,
            "english_row": english_row_idx,
            "header": [c for c in all_rows[header_row_idx] if c is not None],
            "english": [c for c in all_rows[english_row_idx] if c is not None] if english_row_idx is not None else [],
            "data_samples": []
        }
        for row in all_rows[header_row_idx+1:header_row_idx+4]:
            clean = [str(c) if (c is not None and hasattr(c, 'strftime')) else c for c in row]
            details[sheet]["data_samples"].append(clean)

with open('d:/VSCode_Workspace/jetlag_snake/scripts/sheet_details.json', 'w', encoding='utf-8') as f:
    json.dump(details, f, ensure_ascii=False, indent=2, default=str)
