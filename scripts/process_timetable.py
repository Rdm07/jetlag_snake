"""
process_timetable.py
Parses ktx_timetable.xlsx → public/data/timetable.json

Output format:
  { "yongsan_to_daejeon": [{ trainNum, type, depart: "HH:MM", arrive: "HH:MM" }, ...], ... }

Run with:
  C:/Users/droha/anaconda3/envs/env_jetlagsnake/python.exe scripts/process_timetable.py
"""

import openpyxl
import json
import os
import sys

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# ---------------------------------------------------------------------------
# Station name (as it appears in xlsx English rows) → game node ID
# ---------------------------------------------------------------------------
STATION_MAPPING = {
    "Seoul":              "yongsan",
    "Yongsan":            "yongsan",
    "Sangbong":           "sangbong",
    "PyeongtaekJije":     "pyeongtaek",
    "Cheonan-Asan":       "cheonan_asan",
    "Osong":              "osong",
    "Daejeon":            "daejeon",
    "Seodaejeon":         "seodaejeon",
    "Seodaejeon ":        "seodaejeon",   # trailing space in xlsx
    "Iksan":              "iksan",
    "Jeongeup":           "jeongeup",
    "Gwangju Songjeong":  "gwangju_songjeong",
    "Gwangju songjeong":  "gwangju_songjeong",
    "Suncheon":           "suncheon",
    "Gimcheon(Gumi)":     "gimcheon",
    "Dongdaegu":          "dongdaegu",
    "Gyeongju":           "gyeongju",
    "Bujeon":             "bujeon",
    "Seowonju":           "seowonju",
    "Jecheon":            "jecheon",
    "Yeongju":            "yeongju",
    "Yeongcheon":         "yeongcheon",
    "Jinbu(Odaesan)":     "jinbu",
    "Donghae":            "donghae",
    "Jeongdongjin":       "jeongdongjin",
    "Chungju":            "chungju",
}

# All undirected edges in the game graph (from network.ts)
SHOW_EDGES = [
    ("sangbong",    "yongsan"),
    ("sangbong",    "jinbu"),
    ("yongsan",     "pyeongtaek"),
    ("yongsan",     "bongyang"),
    ("pyeongtaek",  "anjung"),
    ("pyeongtaek",  "cheonan"),
    ("cheonan",     "cheonan_asan"),
    ("cheonan",     "osong"),
    ("cheonan_asan","osong"),
    ("osong",       "chungju"),
    ("osong",       "jochiwon"),
    ("chungju",     "jecheon"),
    ("bongyang",    "seowonju"),
    ("seowonju",    "jecheon"),
    ("seowonju",    "dongbaeksan"),
    ("jecheon",     "yeongju"),
    ("jinbu",       "dongbaeksan"),
    ("jinbu",       "jeongdongjin"),
    ("dongbaeksan", "donghae"),
    ("jeongdongjin","donghae"),
    ("jochiwon",    "daejeon"),
    ("jochiwon",    "iksan"),
    ("daejeon",     "seodaejeon"),
    ("seodaejeon",  "iksan"),
    ("iksan",       "hongseong"),
    ("iksan",       "jeongeup"),
    ("jeongeup",    "gwangju_songjeong"),
    ("gwangju_songjeong", "suncheon"),
    ("suncheon",    "samnangjin"),
    ("yeongju",     "gimcheon"),
    ("yeongju",     "yeongcheon"),
    ("gimcheon",    "dongdaegu"),
    ("yeongcheon",  "dongdaegu"),
    ("yeongcheon",  "ahwa"),
    ("dongdaegu",   "seogyeongju"),
    ("seogyeongju", "gyeongju"),
    ("dongdaegu",   "samnangjin"),
    ("samnangjin",  "bujeon"),
    ("bujeon",      "sasang"),
    ("bujeon",      "gyeongju"),
    ("gyeongju",    "donghae"),
]

# Hardcoded local train services for edges with no KTX data.
# Format: "a_to_b": [{ trainNum, type, depart: "HH:MM", arrive: "HH:MM" }, ...]
# Both directions must be listed separately.
# Approx 30-45 min frequency, 07:00–18:30 departures, ~15-35 min travel times.
LOCAL_SERVICES = {
    # yongsan ↔ bongyang  (~40 min, ITX/local)
    "yongsan_to_bongyang": [
        {"trainNum": "L101", "type": "Local", "depart": "07:10", "arrive": "07:50"},
        {"trainNum": "L103", "type": "Local", "depart": "08:10", "arrive": "08:50"},
        {"trainNum": "L105", "type": "Local", "depart": "09:10", "arrive": "09:50"},
        {"trainNum": "L107", "type": "Local", "depart": "10:10", "arrive": "10:50"},
        {"trainNum": "L109", "type": "Local", "depart": "11:10", "arrive": "11:50"},
        {"trainNum": "L111", "type": "Local", "depart": "12:10", "arrive": "12:50"},
        {"trainNum": "L113", "type": "Local", "depart": "13:10", "arrive": "13:50"},
        {"trainNum": "L115", "type": "Local", "depart": "14:10", "arrive": "14:50"},
        {"trainNum": "L117", "type": "Local", "depart": "15:10", "arrive": "15:50"},
        {"trainNum": "L119", "type": "Local", "depart": "16:10", "arrive": "16:50"},
        {"trainNum": "L121", "type": "Local", "depart": "17:10", "arrive": "17:50"},
        {"trainNum": "L123", "type": "Local", "depart": "18:10", "arrive": "18:50"},
    ],
    "bongyang_to_yongsan": [
        {"trainNum": "L102", "type": "Local", "depart": "07:20", "arrive": "08:00"},
        {"trainNum": "L104", "type": "Local", "depart": "08:20", "arrive": "09:00"},
        {"trainNum": "L106", "type": "Local", "depart": "09:20", "arrive": "10:00"},
        {"trainNum": "L108", "type": "Local", "depart": "10:20", "arrive": "11:00"},
        {"trainNum": "L110", "type": "Local", "depart": "11:20", "arrive": "12:00"},
        {"trainNum": "L112", "type": "Local", "depart": "12:20", "arrive": "13:00"},
        {"trainNum": "L114", "type": "Local", "depart": "13:20", "arrive": "14:00"},
        {"trainNum": "L116", "type": "Local", "depart": "14:20", "arrive": "15:00"},
        {"trainNum": "L118", "type": "Local", "depart": "15:20", "arrive": "16:00"},
        {"trainNum": "L120", "type": "Local", "depart": "16:20", "arrive": "17:00"},
        {"trainNum": "L122", "type": "Local", "depart": "17:20", "arrive": "18:00"},
        {"trainNum": "L124", "type": "Local", "depart": "18:20", "arrive": "19:00"},
    ],
    # bongyang ↔ seowonju (~20 min)
    "bongyang_to_seowonju": [
        {"trainNum": "L201", "type": "Local", "depart": "07:05", "arrive": "07:25"},
        {"trainNum": "L203", "type": "Local", "depart": "08:05", "arrive": "08:25"},
        {"trainNum": "L205", "type": "Local", "depart": "09:05", "arrive": "09:25"},
        {"trainNum": "L207", "type": "Local", "depart": "10:05", "arrive": "10:25"},
        {"trainNum": "L209", "type": "Local", "depart": "11:05", "arrive": "11:25"},
        {"trainNum": "L211", "type": "Local", "depart": "12:05", "arrive": "12:25"},
        {"trainNum": "L213", "type": "Local", "depart": "13:05", "arrive": "13:25"},
        {"trainNum": "L215", "type": "Local", "depart": "14:05", "arrive": "14:25"},
        {"trainNum": "L217", "type": "Local", "depart": "15:05", "arrive": "15:25"},
        {"trainNum": "L219", "type": "Local", "depart": "16:05", "arrive": "16:25"},
        {"trainNum": "L221", "type": "Local", "depart": "17:05", "arrive": "17:25"},
        {"trainNum": "L223", "type": "Local", "depart": "18:05", "arrive": "18:25"},
    ],
    "seowonju_to_bongyang": [
        {"trainNum": "L202", "type": "Local", "depart": "07:30", "arrive": "07:50"},
        {"trainNum": "L204", "type": "Local", "depart": "08:30", "arrive": "08:50"},
        {"trainNum": "L206", "type": "Local", "depart": "09:30", "arrive": "09:50"},
        {"trainNum": "L208", "type": "Local", "depart": "10:30", "arrive": "10:50"},
        {"trainNum": "L210", "type": "Local", "depart": "11:30", "arrive": "11:50"},
        {"trainNum": "L212", "type": "Local", "depart": "12:30", "arrive": "12:50"},
        {"trainNum": "L214", "type": "Local", "depart": "13:30", "arrive": "13:50"},
        {"trainNum": "L216", "type": "Local", "depart": "14:30", "arrive": "14:50"},
        {"trainNum": "L218", "type": "Local", "depart": "15:30", "arrive": "15:50"},
        {"trainNum": "L220", "type": "Local", "depart": "16:30", "arrive": "16:50"},
        {"trainNum": "L222", "type": "Local", "depart": "17:30", "arrive": "17:50"},
        {"trainNum": "L224", "type": "Local", "depart": "18:30", "arrive": "18:50"},
    ],
    # pyeongtaek ↔ anjung (~15 min)
    "pyeongtaek_to_anjung": [
        {"trainNum": "L301", "type": "Local", "depart": "07:00", "arrive": "07:15"},
        {"trainNum": "L303", "type": "Local", "depart": "07:45", "arrive": "08:00"},
        {"trainNum": "L305", "type": "Local", "depart": "08:30", "arrive": "08:45"},
        {"trainNum": "L307", "type": "Local", "depart": "09:30", "arrive": "09:45"},
        {"trainNum": "L309", "type": "Local", "depart": "10:30", "arrive": "10:45"},
        {"trainNum": "L311", "type": "Local", "depart": "11:30", "arrive": "11:45"},
        {"trainNum": "L313", "type": "Local", "depart": "12:30", "arrive": "12:45"},
        {"trainNum": "L315", "type": "Local", "depart": "13:30", "arrive": "13:45"},
        {"trainNum": "L317", "type": "Local", "depart": "14:30", "arrive": "14:45"},
        {"trainNum": "L319", "type": "Local", "depart": "15:30", "arrive": "15:45"},
        {"trainNum": "L321", "type": "Local", "depart": "16:30", "arrive": "16:45"},
        {"trainNum": "L323", "type": "Local", "depart": "17:30", "arrive": "17:45"},
        {"trainNum": "L325", "type": "Local", "depart": "18:30", "arrive": "18:45"},
    ],
    "anjung_to_pyeongtaek": [
        {"trainNum": "L302", "type": "Local", "depart": "07:10", "arrive": "07:25"},
        {"trainNum": "L304", "type": "Local", "depart": "07:55", "arrive": "08:10"},
        {"trainNum": "L306", "type": "Local", "depart": "08:55", "arrive": "09:10"},
        {"trainNum": "L308", "type": "Local", "depart": "09:55", "arrive": "10:10"},
        {"trainNum": "L310", "type": "Local", "depart": "10:55", "arrive": "11:10"},
        {"trainNum": "L312", "type": "Local", "depart": "11:55", "arrive": "12:10"},
        {"trainNum": "L314", "type": "Local", "depart": "12:55", "arrive": "13:10"},
        {"trainNum": "L316", "type": "Local", "depart": "13:55", "arrive": "14:10"},
        {"trainNum": "L318", "type": "Local", "depart": "14:55", "arrive": "15:10"},
        {"trainNum": "L320", "type": "Local", "depart": "15:55", "arrive": "16:10"},
        {"trainNum": "L322", "type": "Local", "depart": "16:55", "arrive": "17:10"},
        {"trainNum": "L324", "type": "Local", "depart": "17:55", "arrive": "18:10"},
        {"trainNum": "L326", "type": "Local", "depart": "18:40", "arrive": "18:55"},
    ],
    # pyeongtaek ↔ cheonan (~20 min)
    "pyeongtaek_to_cheonan": [
        {"trainNum": "L401", "type": "Local", "depart": "07:10", "arrive": "07:30"},
        {"trainNum": "L403", "type": "Local", "depart": "08:10", "arrive": "08:30"},
        {"trainNum": "L405", "type": "Local", "depart": "09:10", "arrive": "09:30"},
        {"trainNum": "L407", "type": "Local", "depart": "10:10", "arrive": "10:30"},
        {"trainNum": "L409", "type": "Local", "depart": "11:10", "arrive": "11:30"},
        {"trainNum": "L411", "type": "Local", "depart": "12:10", "arrive": "12:30"},
        {"trainNum": "L413", "type": "Local", "depart": "13:10", "arrive": "13:30"},
        {"trainNum": "L415", "type": "Local", "depart": "14:10", "arrive": "14:30"},
        {"trainNum": "L417", "type": "Local", "depart": "15:10", "arrive": "15:30"},
        {"trainNum": "L419", "type": "Local", "depart": "16:10", "arrive": "16:30"},
        {"trainNum": "L421", "type": "Local", "depart": "17:10", "arrive": "17:30"},
        {"trainNum": "L423", "type": "Local", "depart": "18:10", "arrive": "18:30"},
    ],
    "cheonan_to_pyeongtaek": [
        {"trainNum": "L402", "type": "Local", "depart": "07:20", "arrive": "07:40"},
        {"trainNum": "L404", "type": "Local", "depart": "08:20", "arrive": "08:40"},
        {"trainNum": "L406", "type": "Local", "depart": "09:20", "arrive": "09:40"},
        {"trainNum": "L408", "type": "Local", "depart": "10:20", "arrive": "10:40"},
        {"trainNum": "L410", "type": "Local", "depart": "11:20", "arrive": "11:40"},
        {"trainNum": "L412", "type": "Local", "depart": "12:20", "arrive": "12:40"},
        {"trainNum": "L414", "type": "Local", "depart": "13:20", "arrive": "13:40"},
        {"trainNum": "L416", "type": "Local", "depart": "14:20", "arrive": "14:40"},
        {"trainNum": "L418", "type": "Local", "depart": "15:20", "arrive": "15:40"},
        {"trainNum": "L420", "type": "Local", "depart": "16:20", "arrive": "16:40"},
        {"trainNum": "L422", "type": "Local", "depart": "17:20", "arrive": "17:40"},
        {"trainNum": "L424", "type": "Local", "depart": "18:20", "arrive": "18:40"},
    ],
    # cheonan ↔ cheonan_asan (~10 min)
    "cheonan_to_cheonan_asan": [
        {"trainNum": "L501", "type": "Local", "depart": "07:05", "arrive": "07:15"},
        {"trainNum": "L503", "type": "Local", "depart": "07:45", "arrive": "07:55"},
        {"trainNum": "L505", "type": "Local", "depart": "08:35", "arrive": "08:45"},
        {"trainNum": "L507", "type": "Local", "depart": "09:35", "arrive": "09:45"},
        {"trainNum": "L509", "type": "Local", "depart": "10:35", "arrive": "10:45"},
        {"trainNum": "L511", "type": "Local", "depart": "11:35", "arrive": "11:45"},
        {"trainNum": "L513", "type": "Local", "depart": "12:35", "arrive": "12:45"},
        {"trainNum": "L515", "type": "Local", "depart": "13:35", "arrive": "13:45"},
        {"trainNum": "L517", "type": "Local", "depart": "14:35", "arrive": "14:45"},
        {"trainNum": "L519", "type": "Local", "depart": "15:35", "arrive": "15:45"},
        {"trainNum": "L521", "type": "Local", "depart": "16:35", "arrive": "16:45"},
        {"trainNum": "L523", "type": "Local", "depart": "17:35", "arrive": "17:45"},
        {"trainNum": "L525", "type": "Local", "depart": "18:35", "arrive": "18:45"},
    ],
    "cheonan_asan_to_cheonan": [
        {"trainNum": "L502", "type": "Local", "depart": "07:10", "arrive": "07:20"},
        {"trainNum": "L504", "type": "Local", "depart": "07:50", "arrive": "08:00"},
        {"trainNum": "L506", "type": "Local", "depart": "08:50", "arrive": "09:00"},
        {"trainNum": "L508", "type": "Local", "depart": "09:50", "arrive": "10:00"},
        {"trainNum": "L510", "type": "Local", "depart": "10:50", "arrive": "11:00"},
        {"trainNum": "L512", "type": "Local", "depart": "11:50", "arrive": "12:00"},
        {"trainNum": "L514", "type": "Local", "depart": "12:50", "arrive": "13:00"},
        {"trainNum": "L516", "type": "Local", "depart": "13:50", "arrive": "14:00"},
        {"trainNum": "L518", "type": "Local", "depart": "14:50", "arrive": "15:00"},
        {"trainNum": "L520", "type": "Local", "depart": "15:50", "arrive": "16:00"},
        {"trainNum": "L522", "type": "Local", "depart": "16:50", "arrive": "17:00"},
        {"trainNum": "L524", "type": "Local", "depart": "17:50", "arrive": "18:00"},
        {"trainNum": "L526", "type": "Local", "depart": "18:40", "arrive": "18:50"},
    ],
    # cheonan ↔ osong (~15 min, local)
    "cheonan_to_osong": [
        {"trainNum": "L601", "type": "Local", "depart": "07:10", "arrive": "07:25"},
        {"trainNum": "L603", "type": "Local", "depart": "08:10", "arrive": "08:25"},
        {"trainNum": "L605", "type": "Local", "depart": "09:10", "arrive": "09:25"},
        {"trainNum": "L607", "type": "Local", "depart": "10:10", "arrive": "10:25"},
        {"trainNum": "L609", "type": "Local", "depart": "11:10", "arrive": "11:25"},
        {"trainNum": "L611", "type": "Local", "depart": "12:10", "arrive": "12:25"},
        {"trainNum": "L613", "type": "Local", "depart": "13:10", "arrive": "13:25"},
        {"trainNum": "L615", "type": "Local", "depart": "14:10", "arrive": "14:25"},
        {"trainNum": "L617", "type": "Local", "depart": "15:10", "arrive": "15:25"},
        {"trainNum": "L619", "type": "Local", "depart": "16:10", "arrive": "16:25"},
        {"trainNum": "L621", "type": "Local", "depart": "17:10", "arrive": "17:25"},
        {"trainNum": "L623", "type": "Local", "depart": "18:10", "arrive": "18:25"},
    ],
    "osong_to_cheonan": [
        {"trainNum": "L602", "type": "Local", "depart": "07:20", "arrive": "07:35"},
        {"trainNum": "L604", "type": "Local", "depart": "08:20", "arrive": "08:35"},
        {"trainNum": "L606", "type": "Local", "depart": "09:20", "arrive": "09:35"},
        {"trainNum": "L608", "type": "Local", "depart": "10:20", "arrive": "10:35"},
        {"trainNum": "L610", "type": "Local", "depart": "11:20", "arrive": "11:35"},
        {"trainNum": "L612", "type": "Local", "depart": "12:20", "arrive": "12:35"},
        {"trainNum": "L614", "type": "Local", "depart": "13:20", "arrive": "13:35"},
        {"trainNum": "L616", "type": "Local", "depart": "14:20", "arrive": "14:35"},
        {"trainNum": "L618", "type": "Local", "depart": "15:20", "arrive": "15:35"},
        {"trainNum": "L620", "type": "Local", "depart": "16:20", "arrive": "16:35"},
        {"trainNum": "L622", "type": "Local", "depart": "17:20", "arrive": "17:35"},
        {"trainNum": "L624", "type": "Local", "depart": "18:20", "arrive": "18:35"},
    ],
    # osong ↔ jochiwon (~20 min, local)
    "osong_to_jochiwon": [
        {"trainNum": "L701", "type": "Local", "depart": "07:15", "arrive": "07:35"},
        {"trainNum": "L703", "type": "Local", "depart": "08:45", "arrive": "09:05"},
        {"trainNum": "L705", "type": "Local", "depart": "10:15", "arrive": "10:35"},
        {"trainNum": "L707", "type": "Local", "depart": "11:45", "arrive": "12:05"},
        {"trainNum": "L709", "type": "Local", "depart": "13:15", "arrive": "13:35"},
        {"trainNum": "L711", "type": "Local", "depart": "14:45", "arrive": "15:05"},
        {"trainNum": "L713", "type": "Local", "depart": "16:15", "arrive": "16:35"},
        {"trainNum": "L715", "type": "Local", "depart": "17:45", "arrive": "18:05"},
    ],
    "jochiwon_to_osong": [
        {"trainNum": "L702", "type": "Local", "depart": "07:50", "arrive": "08:10"},
        {"trainNum": "L704", "type": "Local", "depart": "09:20", "arrive": "09:40"},
        {"trainNum": "L706", "type": "Local", "depart": "10:50", "arrive": "11:10"},
        {"trainNum": "L708", "type": "Local", "depart": "12:20", "arrive": "12:40"},
        {"trainNum": "L710", "type": "Local", "depart": "13:50", "arrive": "14:10"},
        {"trainNum": "L712", "type": "Local", "depart": "15:20", "arrive": "15:40"},
        {"trainNum": "L714", "type": "Local", "depart": "16:50", "arrive": "17:10"},
        {"trainNum": "L716", "type": "Local", "depart": "18:20", "arrive": "18:40"},
    ],
    # jochiwon ↔ daejeon (~18 min, local)
    "jochiwon_to_daejeon": [
        {"trainNum": "L801", "type": "Local", "depart": "07:10", "arrive": "07:28"},
        {"trainNum": "L803", "type": "Local", "depart": "08:40", "arrive": "08:58"},
        {"trainNum": "L805", "type": "Local", "depart": "10:10", "arrive": "10:28"},
        {"trainNum": "L807", "type": "Local", "depart": "11:40", "arrive": "11:58"},
        {"trainNum": "L809", "type": "Local", "depart": "13:10", "arrive": "13:28"},
        {"trainNum": "L811", "type": "Local", "depart": "14:40", "arrive": "14:58"},
        {"trainNum": "L813", "type": "Local", "depart": "16:10", "arrive": "16:28"},
        {"trainNum": "L815", "type": "Local", "depart": "17:40", "arrive": "17:58"},
    ],
    "daejeon_to_jochiwon": [
        {"trainNum": "L802", "type": "Local", "depart": "07:35", "arrive": "07:53"},
        {"trainNum": "L804", "type": "Local", "depart": "09:05", "arrive": "09:23"},
        {"trainNum": "L806", "type": "Local", "depart": "10:35", "arrive": "10:53"},
        {"trainNum": "L808", "type": "Local", "depart": "12:05", "arrive": "12:23"},
        {"trainNum": "L810", "type": "Local", "depart": "13:35", "arrive": "13:53"},
        {"trainNum": "L812", "type": "Local", "depart": "15:05", "arrive": "15:23"},
        {"trainNum": "L814", "type": "Local", "depart": "16:35", "arrive": "16:53"},
        {"trainNum": "L816", "type": "Local", "depart": "18:05", "arrive": "18:23"},
    ],
    # jochiwon ↔ iksan (~30 min, local)
    "jochiwon_to_iksan": [
        {"trainNum": "L901", "type": "Local", "depart": "07:20", "arrive": "07:50"},
        {"trainNum": "L903", "type": "Local", "depart": "09:20", "arrive": "09:50"},
        {"trainNum": "L905", "type": "Local", "depart": "11:20", "arrive": "11:50"},
        {"trainNum": "L907", "type": "Local", "depart": "13:20", "arrive": "13:50"},
        {"trainNum": "L909", "type": "Local", "depart": "15:20", "arrive": "15:50"},
        {"trainNum": "L911", "type": "Local", "depart": "17:20", "arrive": "17:50"},
    ],
    "iksan_to_jochiwon": [
        {"trainNum": "L902", "type": "Local", "depart": "07:55", "arrive": "08:25"},
        {"trainNum": "L904", "type": "Local", "depart": "09:55", "arrive": "10:25"},
        {"trainNum": "L906", "type": "Local", "depart": "11:55", "arrive": "12:25"},
        {"trainNum": "L908", "type": "Local", "depart": "13:55", "arrive": "14:25"},
        {"trainNum": "L910", "type": "Local", "depart": "15:55", "arrive": "16:25"},
        {"trainNum": "L912", "type": "Local", "depart": "17:55", "arrive": "18:25"},
    ],
    # daejeon ↔ seodaejeon (~10 min, very short)
    "daejeon_to_seodaejeon": [
        {"trainNum": "L1001", "type": "Local", "depart": "07:05", "arrive": "07:15"},
        {"trainNum": "L1003", "type": "Local", "depart": "07:45", "arrive": "07:55"},
        {"trainNum": "L1005", "type": "Local", "depart": "08:45", "arrive": "08:55"},
        {"trainNum": "L1007", "type": "Local", "depart": "09:45", "arrive": "09:55"},
        {"trainNum": "L1009", "type": "Local", "depart": "10:45", "arrive": "10:55"},
        {"trainNum": "L1011", "type": "Local", "depart": "11:45", "arrive": "11:55"},
        {"trainNum": "L1013", "type": "Local", "depart": "12:45", "arrive": "12:55"},
        {"trainNum": "L1015", "type": "Local", "depart": "13:45", "arrive": "13:55"},
        {"trainNum": "L1017", "type": "Local", "depart": "14:45", "arrive": "14:55"},
        {"trainNum": "L1019", "type": "Local", "depart": "15:45", "arrive": "15:55"},
        {"trainNum": "L1021", "type": "Local", "depart": "16:45", "arrive": "16:55"},
        {"trainNum": "L1023", "type": "Local", "depart": "17:45", "arrive": "17:55"},
        {"trainNum": "L1025", "type": "Local", "depart": "18:40", "arrive": "18:50"},
    ],
    "seodaejeon_to_daejeon": [
        {"trainNum": "L1002", "type": "Local", "depart": "07:10", "arrive": "07:20"},
        {"trainNum": "L1004", "type": "Local", "depart": "07:50", "arrive": "08:00"},
        {"trainNum": "L1006", "type": "Local", "depart": "08:50", "arrive": "09:00"},
        {"trainNum": "L1008", "type": "Local", "depart": "09:50", "arrive": "10:00"},
        {"trainNum": "L1010", "type": "Local", "depart": "10:50", "arrive": "11:00"},
        {"trainNum": "L1012", "type": "Local", "depart": "11:50", "arrive": "12:00"},
        {"trainNum": "L1014", "type": "Local", "depart": "12:50", "arrive": "13:00"},
        {"trainNum": "L1016", "type": "Local", "depart": "13:50", "arrive": "14:00"},
        {"trainNum": "L1018", "type": "Local", "depart": "14:50", "arrive": "15:00"},
        {"trainNum": "L1020", "type": "Local", "depart": "15:50", "arrive": "16:00"},
        {"trainNum": "L1022", "type": "Local", "depart": "16:50", "arrive": "17:00"},
        {"trainNum": "L1024", "type": "Local", "depart": "17:50", "arrive": "18:00"},
        {"trainNum": "L1026", "type": "Local", "depart": "18:45", "arrive": "18:55"},
    ],
    # iksan ↔ hongseong (~35 min, local)
    "iksan_to_hongseong": [
        {"trainNum": "L1101", "type": "Local", "depart": "07:10", "arrive": "07:45"},
        {"trainNum": "L1103", "type": "Local", "depart": "09:10", "arrive": "09:45"},
        {"trainNum": "L1105", "type": "Local", "depart": "11:10", "arrive": "11:45"},
        {"trainNum": "L1107", "type": "Local", "depart": "13:10", "arrive": "13:45"},
        {"trainNum": "L1109", "type": "Local", "depart": "15:10", "arrive": "15:45"},
        {"trainNum": "L1111", "type": "Local", "depart": "17:10", "arrive": "17:45"},
    ],
    "hongseong_to_iksan": [
        {"trainNum": "L1102", "type": "Local", "depart": "07:50", "arrive": "08:25"},
        {"trainNum": "L1104", "type": "Local", "depart": "09:50", "arrive": "10:25"},
        {"trainNum": "L1106", "type": "Local", "depart": "11:50", "arrive": "12:25"},
        {"trainNum": "L1108", "type": "Local", "depart": "13:50", "arrive": "14:25"},
        {"trainNum": "L1110", "type": "Local", "depart": "15:50", "arrive": "16:25"},
        {"trainNum": "L1112", "type": "Local", "depart": "17:50", "arrive": "18:25"},
    ],
    # osong ↔ chungju (~40 min — connects via the Jungbu Inland line area, local bus/train)
    "osong_to_chungju": [
        {"trainNum": "L1201", "type": "Local", "depart": "07:20", "arrive": "08:00"},
        {"trainNum": "L1203", "type": "Local", "depart": "09:20", "arrive": "10:00"},
        {"trainNum": "L1205", "type": "Local", "depart": "11:20", "arrive": "12:00"},
        {"trainNum": "L1207", "type": "Local", "depart": "13:20", "arrive": "14:00"},
        {"trainNum": "L1209", "type": "Local", "depart": "15:20", "arrive": "16:00"},
        {"trainNum": "L1211", "type": "Local", "depart": "17:20", "arrive": "18:00"},
    ],
    "chungju_to_osong": [
        {"trainNum": "L1202", "type": "Local", "depart": "07:05", "arrive": "07:45"},
        {"trainNum": "L1204", "type": "Local", "depart": "09:05", "arrive": "09:45"},
        {"trainNum": "L1206", "type": "Local", "depart": "11:05", "arrive": "11:45"},
        {"trainNum": "L1208", "type": "Local", "depart": "13:05", "arrive": "13:45"},
        {"trainNum": "L1210", "type": "Local", "depart": "15:05", "arrive": "15:45"},
        {"trainNum": "L1212", "type": "Local", "depart": "17:05", "arrive": "17:45"},
    ],
    # seowonju ↔ dongbaeksan (~25 min, local)
    "seowonju_to_dongbaeksan": [
        {"trainNum": "L1301", "type": "Local", "depart": "07:30", "arrive": "07:55"},
        {"trainNum": "L1303", "type": "Local", "depart": "09:30", "arrive": "09:55"},
        {"trainNum": "L1305", "type": "Local", "depart": "11:30", "arrive": "11:55"},
        {"trainNum": "L1307", "type": "Local", "depart": "13:30", "arrive": "13:55"},
        {"trainNum": "L1309", "type": "Local", "depart": "15:30", "arrive": "15:55"},
        {"trainNum": "L1311", "type": "Local", "depart": "17:30", "arrive": "17:55"},
    ],
    "dongbaeksan_to_seowonju": [
        {"trainNum": "L1302", "type": "Local", "depart": "08:05", "arrive": "08:30"},
        {"trainNum": "L1304", "type": "Local", "depart": "10:05", "arrive": "10:30"},
        {"trainNum": "L1306", "type": "Local", "depart": "12:05", "arrive": "12:30"},
        {"trainNum": "L1308", "type": "Local", "depart": "14:05", "arrive": "14:30"},
        {"trainNum": "L1310", "type": "Local", "depart": "16:05", "arrive": "16:30"},
        {"trainNum": "L1312", "type": "Local", "depart": "18:05", "arrive": "18:30"},
    ],
    # jinbu ↔ dongbaeksan (~30 min, local)
    "jinbu_to_dongbaeksan": [
        {"trainNum": "L1401", "type": "Local", "depart": "07:15", "arrive": "07:45"},
        {"trainNum": "L1403", "type": "Local", "depart": "09:15", "arrive": "09:45"},
        {"trainNum": "L1405", "type": "Local", "depart": "11:15", "arrive": "11:45"},
        {"trainNum": "L1407", "type": "Local", "depart": "13:15", "arrive": "13:45"},
        {"trainNum": "L1409", "type": "Local", "depart": "15:15", "arrive": "15:45"},
        {"trainNum": "L1411", "type": "Local", "depart": "17:15", "arrive": "17:45"},
    ],
    "dongbaeksan_to_jinbu": [
        {"trainNum": "L1402", "type": "Local", "depart": "07:50", "arrive": "08:20"},
        {"trainNum": "L1404", "type": "Local", "depart": "09:50", "arrive": "10:20"},
        {"trainNum": "L1406", "type": "Local", "depart": "11:50", "arrive": "12:20"},
        {"trainNum": "L1408", "type": "Local", "depart": "13:50", "arrive": "14:20"},
        {"trainNum": "L1410", "type": "Local", "depart": "15:50", "arrive": "16:20"},
        {"trainNum": "L1412", "type": "Local", "depart": "17:50", "arrive": "18:20"},
    ],
    # dongbaeksan ↔ donghae (~20 min, local)
    "dongbaeksan_to_donghae": [
        {"trainNum": "L1501", "type": "Local", "depart": "07:20", "arrive": "07:40"},
        {"trainNum": "L1503", "type": "Local", "depart": "09:20", "arrive": "09:40"},
        {"trainNum": "L1505", "type": "Local", "depart": "11:20", "arrive": "11:40"},
        {"trainNum": "L1507", "type": "Local", "depart": "13:20", "arrive": "13:40"},
        {"trainNum": "L1509", "type": "Local", "depart": "15:20", "arrive": "15:40"},
        {"trainNum": "L1511", "type": "Local", "depart": "17:20", "arrive": "17:40"},
    ],
    "donghae_to_dongbaeksan": [
        {"trainNum": "L1502", "type": "Local", "depart": "07:45", "arrive": "08:05"},
        {"trainNum": "L1504", "type": "Local", "depart": "09:45", "arrive": "10:05"},
        {"trainNum": "L1506", "type": "Local", "depart": "11:45", "arrive": "12:05"},
        {"trainNum": "L1508", "type": "Local", "depart": "13:45", "arrive": "14:05"},
        {"trainNum": "L1510", "type": "Local", "depart": "15:45", "arrive": "16:05"},
        {"trainNum": "L1512", "type": "Local", "depart": "17:45", "arrive": "18:05"},
    ],
    # yeongju ↔ gimcheon (~35 min, local — connects Jungang line to Gyeongbu line)
    "yeongju_to_gimcheon": [
        {"trainNum": "L1601", "type": "Local", "depart": "07:10", "arrive": "07:45"},
        {"trainNum": "L1603", "type": "Local", "depart": "09:10", "arrive": "09:45"},
        {"trainNum": "L1605", "type": "Local", "depart": "11:10", "arrive": "11:45"},
        {"trainNum": "L1607", "type": "Local", "depart": "13:10", "arrive": "13:45"},
        {"trainNum": "L1609", "type": "Local", "depart": "15:10", "arrive": "15:45"},
        {"trainNum": "L1611", "type": "Local", "depart": "17:10", "arrive": "17:45"},
    ],
    "gimcheon_to_yeongju": [
        {"trainNum": "L1602", "type": "Local", "depart": "07:50", "arrive": "08:25"},
        {"trainNum": "L1604", "type": "Local", "depart": "09:50", "arrive": "10:25"},
        {"trainNum": "L1606", "type": "Local", "depart": "11:50", "arrive": "12:25"},
        {"trainNum": "L1608", "type": "Local", "depart": "13:50", "arrive": "14:25"},
        {"trainNum": "L1610", "type": "Local", "depart": "15:50", "arrive": "16:25"},
        {"trainNum": "L1612", "type": "Local", "depart": "17:50", "arrive": "18:25"},
    ],
    # dongdaegu ↔ seogyeongju (~25 min, local)
    "dongdaegu_to_seogyeongju": [
        {"trainNum": "L1701", "type": "Local", "depart": "07:10", "arrive": "07:35"},
        {"trainNum": "L1703", "type": "Local", "depart": "08:40", "arrive": "09:05"},
        {"trainNum": "L1705", "type": "Local", "depart": "10:10", "arrive": "10:35"},
        {"trainNum": "L1707", "type": "Local", "depart": "11:40", "arrive": "12:05"},
        {"trainNum": "L1709", "type": "Local", "depart": "13:10", "arrive": "13:35"},
        {"trainNum": "L1711", "type": "Local", "depart": "14:40", "arrive": "15:05"},
        {"trainNum": "L1713", "type": "Local", "depart": "16:10", "arrive": "16:35"},
        {"trainNum": "L1715", "type": "Local", "depart": "17:40", "arrive": "18:05"},
    ],
    "seogyeongju_to_dongdaegu": [
        {"trainNum": "L1702", "type": "Local", "depart": "07:40", "arrive": "08:05"},
        {"trainNum": "L1704", "type": "Local", "depart": "09:10", "arrive": "09:35"},
        {"trainNum": "L1706", "type": "Local", "depart": "10:40", "arrive": "11:05"},
        {"trainNum": "L1708", "type": "Local", "depart": "12:10", "arrive": "12:35"},
        {"trainNum": "L1710", "type": "Local", "depart": "13:40", "arrive": "14:05"},
        {"trainNum": "L1712", "type": "Local", "depart": "15:10", "arrive": "15:35"},
        {"trainNum": "L1714", "type": "Local", "depart": "16:40", "arrive": "17:05"},
        {"trainNum": "L1716", "type": "Local", "depart": "18:10", "arrive": "18:35"},
    ],
    # seogyeongju ↔ gyeongju (~15 min, local)
    "seogyeongju_to_gyeongju": [
        {"trainNum": "L1801", "type": "Local", "depart": "07:05", "arrive": "07:20"},
        {"trainNum": "L1803", "type": "Local", "depart": "08:05", "arrive": "08:20"},
        {"trainNum": "L1805", "type": "Local", "depart": "09:05", "arrive": "09:20"},
        {"trainNum": "L1807", "type": "Local", "depart": "10:05", "arrive": "10:20"},
        {"trainNum": "L1809", "type": "Local", "depart": "11:05", "arrive": "11:20"},
        {"trainNum": "L1811", "type": "Local", "depart": "12:05", "arrive": "12:20"},
        {"trainNum": "L1813", "type": "Local", "depart": "13:05", "arrive": "13:20"},
        {"trainNum": "L1815", "type": "Local", "depart": "14:05", "arrive": "14:20"},
        {"trainNum": "L1817", "type": "Local", "depart": "15:05", "arrive": "15:20"},
        {"trainNum": "L1819", "type": "Local", "depart": "16:05", "arrive": "16:20"},
        {"trainNum": "L1821", "type": "Local", "depart": "17:05", "arrive": "17:20"},
        {"trainNum": "L1823", "type": "Local", "depart": "18:05", "arrive": "18:20"},
    ],
    "gyeongju_to_seogyeongju": [
        {"trainNum": "L1802", "type": "Local", "depart": "07:10", "arrive": "07:25"},
        {"trainNum": "L1804", "type": "Local", "depart": "08:10", "arrive": "08:25"},
        {"trainNum": "L1806", "type": "Local", "depart": "09:10", "arrive": "09:25"},
        {"trainNum": "L1808", "type": "Local", "depart": "10:10", "arrive": "10:25"},
        {"trainNum": "L1810", "type": "Local", "depart": "11:10", "arrive": "11:25"},
        {"trainNum": "L1812", "type": "Local", "depart": "12:10", "arrive": "12:25"},
        {"trainNum": "L1814", "type": "Local", "depart": "13:10", "arrive": "13:25"},
        {"trainNum": "L1816", "type": "Local", "depart": "14:10", "arrive": "14:25"},
        {"trainNum": "L1818", "type": "Local", "depart": "15:10", "arrive": "15:25"},
        {"trainNum": "L1820", "type": "Local", "depart": "16:10", "arrive": "16:25"},
        {"trainNum": "L1822", "type": "Local", "depart": "17:10", "arrive": "17:25"},
        {"trainNum": "L1824", "type": "Local", "depart": "18:10", "arrive": "18:25"},
    ],
    # bujeon ↔ sasang (~12 min, local)
    "bujeon_to_sasang": [
        {"trainNum": "L1901", "type": "Local", "depart": "07:05", "arrive": "07:17"},
        {"trainNum": "L1903", "type": "Local", "depart": "08:05", "arrive": "08:17"},
        {"trainNum": "L1905", "type": "Local", "depart": "09:05", "arrive": "09:17"},
        {"trainNum": "L1907", "type": "Local", "depart": "10:05", "arrive": "10:17"},
        {"trainNum": "L1909", "type": "Local", "depart": "11:05", "arrive": "11:17"},
        {"trainNum": "L1911", "type": "Local", "depart": "12:05", "arrive": "12:17"},
        {"trainNum": "L1913", "type": "Local", "depart": "13:05", "arrive": "13:17"},
        {"trainNum": "L1915", "type": "Local", "depart": "14:05", "arrive": "14:17"},
        {"trainNum": "L1917", "type": "Local", "depart": "15:05", "arrive": "15:17"},
        {"trainNum": "L1919", "type": "Local", "depart": "16:05", "arrive": "16:17"},
        {"trainNum": "L1921", "type": "Local", "depart": "17:05", "arrive": "17:17"},
        {"trainNum": "L1923", "type": "Local", "depart": "18:05", "arrive": "18:17"},
    ],
    "sasang_to_bujeon": [
        {"trainNum": "L1902", "type": "Local", "depart": "07:10", "arrive": "07:22"},
        {"trainNum": "L1904", "type": "Local", "depart": "08:10", "arrive": "08:22"},
        {"trainNum": "L1906", "type": "Local", "depart": "09:10", "arrive": "09:22"},
        {"trainNum": "L1908", "type": "Local", "depart": "10:10", "arrive": "10:22"},
        {"trainNum": "L1910", "type": "Local", "depart": "11:10", "arrive": "11:22"},
        {"trainNum": "L1912", "type": "Local", "depart": "12:10", "arrive": "12:22"},
        {"trainNum": "L1914", "type": "Local", "depart": "13:10", "arrive": "13:22"},
        {"trainNum": "L1916", "type": "Local", "depart": "14:10", "arrive": "14:22"},
        {"trainNum": "L1918", "type": "Local", "depart": "15:10", "arrive": "15:22"},
        {"trainNum": "L1920", "type": "Local", "depart": "16:10", "arrive": "16:22"},
        {"trainNum": "L1922", "type": "Local", "depart": "17:10", "arrive": "17:22"},
        {"trainNum": "L1924", "type": "Local", "depart": "18:10", "arrive": "18:22"},
    ],
    # suncheon ↔ samnangjin (~55 min, local)
    "suncheon_to_samnangjin": [
        {"trainNum": "L2001", "type": "Local", "depart": "07:10", "arrive": "08:05"},
        {"trainNum": "L2003", "type": "Local", "depart": "09:10", "arrive": "10:05"},
        {"trainNum": "L2005", "type": "Local", "depart": "11:10", "arrive": "12:05"},
        {"trainNum": "L2007", "type": "Local", "depart": "13:10", "arrive": "14:05"},
        {"trainNum": "L2009", "type": "Local", "depart": "15:10", "arrive": "16:05"},
        {"trainNum": "L2011", "type": "Local", "depart": "17:10", "arrive": "18:05"},
    ],
    "samnangjin_to_suncheon": [
        {"trainNum": "L2002", "type": "Local", "depart": "07:15", "arrive": "08:10"},
        {"trainNum": "L2004", "type": "Local", "depart": "09:15", "arrive": "10:10"},
        {"trainNum": "L2006", "type": "Local", "depart": "11:15", "arrive": "12:10"},
        {"trainNum": "L2008", "type": "Local", "depart": "13:15", "arrive": "14:10"},
        {"trainNum": "L2010", "type": "Local", "depart": "15:15", "arrive": "16:10"},
        {"trainNum": "L2012", "type": "Local", "depart": "17:15", "arrive": "18:10"},
    ],
    # yongsan ↔ pyeongtaek (~35 min — most KTX skip PyeongtaekJije; use local/SRT)
    "yongsan_to_pyeongtaek": [
        {"trainNum": "L2201", "type": "Local", "depart": "07:10", "arrive": "07:45"},
        {"trainNum": "L2203", "type": "Local", "depart": "08:10", "arrive": "08:45"},
        {"trainNum": "L2205", "type": "Local", "depart": "09:10", "arrive": "09:45"},
        {"trainNum": "L2207", "type": "Local", "depart": "10:10", "arrive": "10:45"},
        {"trainNum": "L2209", "type": "Local", "depart": "11:10", "arrive": "11:45"},
        {"trainNum": "L2211", "type": "Local", "depart": "12:10", "arrive": "12:45"},
        {"trainNum": "L2213", "type": "Local", "depart": "13:10", "arrive": "13:45"},
        {"trainNum": "L2215", "type": "Local", "depart": "14:10", "arrive": "14:45"},
        {"trainNum": "L2217", "type": "Local", "depart": "15:10", "arrive": "15:45"},
        {"trainNum": "L2219", "type": "Local", "depart": "16:10", "arrive": "16:45"},
        {"trainNum": "L2221", "type": "Local", "depart": "17:10", "arrive": "17:45"},
        {"trainNum": "L2223", "type": "Local", "depart": "18:10", "arrive": "18:45"},
    ],
    "pyeongtaek_to_yongsan": [
        {"trainNum": "L2202", "type": "Local", "depart": "07:20", "arrive": "07:55"},
        {"trainNum": "L2204", "type": "Local", "depart": "08:20", "arrive": "08:55"},
        {"trainNum": "L2206", "type": "Local", "depart": "09:20", "arrive": "09:55"},
        {"trainNum": "L2208", "type": "Local", "depart": "10:20", "arrive": "10:55"},
        {"trainNum": "L2210", "type": "Local", "depart": "11:20", "arrive": "11:55"},
        {"trainNum": "L2212", "type": "Local", "depart": "12:20", "arrive": "12:55"},
        {"trainNum": "L2214", "type": "Local", "depart": "13:20", "arrive": "13:55"},
        {"trainNum": "L2216", "type": "Local", "depart": "14:20", "arrive": "14:55"},
        {"trainNum": "L2218", "type": "Local", "depart": "15:20", "arrive": "15:55"},
        {"trainNum": "L2220", "type": "Local", "depart": "16:20", "arrive": "16:55"},
        {"trainNum": "L2222", "type": "Local", "depart": "17:20", "arrive": "17:55"},
        {"trainNum": "L2224", "type": "Local", "depart": "18:20", "arrive": "18:55"},
    ],
    # sangbong ↔ yongsan (~30 min, local — KTX 강릉선 runs express Seoul–Seowonju)
    "sangbong_to_yongsan": [
        {"trainNum": "L2301", "type": "Local", "depart": "07:05", "arrive": "07:35"},
        {"trainNum": "L2303", "type": "Local", "depart": "08:05", "arrive": "08:35"},
        {"trainNum": "L2305", "type": "Local", "depart": "09:05", "arrive": "09:35"},
        {"trainNum": "L2307", "type": "Local", "depart": "10:05", "arrive": "10:35"},
        {"trainNum": "L2309", "type": "Local", "depart": "11:05", "arrive": "11:35"},
        {"trainNum": "L2311", "type": "Local", "depart": "12:05", "arrive": "12:35"},
        {"trainNum": "L2313", "type": "Local", "depart": "13:05", "arrive": "13:35"},
        {"trainNum": "L2315", "type": "Local", "depart": "14:05", "arrive": "14:35"},
        {"trainNum": "L2317", "type": "Local", "depart": "15:05", "arrive": "15:35"},
        {"trainNum": "L2319", "type": "Local", "depart": "16:05", "arrive": "16:35"},
        {"trainNum": "L2321", "type": "Local", "depart": "17:05", "arrive": "17:35"},
        {"trainNum": "L2323", "type": "Local", "depart": "18:05", "arrive": "18:35"},
    ],
    "yongsan_to_sangbong": [
        {"trainNum": "L2302", "type": "Local", "depart": "07:10", "arrive": "07:40"},
        {"trainNum": "L2304", "type": "Local", "depart": "08:10", "arrive": "08:40"},
        {"trainNum": "L2306", "type": "Local", "depart": "09:10", "arrive": "09:40"},
        {"trainNum": "L2308", "type": "Local", "depart": "10:10", "arrive": "10:40"},
        {"trainNum": "L2310", "type": "Local", "depart": "11:10", "arrive": "11:40"},
        {"trainNum": "L2312", "type": "Local", "depart": "12:10", "arrive": "12:40"},
        {"trainNum": "L2314", "type": "Local", "depart": "13:10", "arrive": "13:40"},
        {"trainNum": "L2316", "type": "Local", "depart": "14:10", "arrive": "14:40"},
        {"trainNum": "L2318", "type": "Local", "depart": "15:10", "arrive": "15:40"},
        {"trainNum": "L2320", "type": "Local", "depart": "16:10", "arrive": "16:40"},
        {"trainNum": "L2322", "type": "Local", "depart": "17:10", "arrive": "17:40"},
        {"trainNum": "L2324", "type": "Local", "depart": "18:10", "arrive": "18:40"},
    ],
    # sangbong ↔ jinbu (~55 min — KTX trains often skip Jinbu; use local/slow KTX)
    "sangbong_to_jinbu": [
        {"trainNum": "L2401", "type": "Local", "depart": "07:15", "arrive": "08:10"},
        {"trainNum": "L2403", "type": "Local", "depart": "09:15", "arrive": "10:10"},
        {"trainNum": "L2405", "type": "Local", "depart": "11:15", "arrive": "12:10"},
        {"trainNum": "L2407", "type": "Local", "depart": "13:15", "arrive": "14:10"},
        {"trainNum": "L2409", "type": "Local", "depart": "15:15", "arrive": "16:10"},
        {"trainNum": "L2411", "type": "Local", "depart": "17:15", "arrive": "18:10"},
    ],
    "jinbu_to_sangbong": [
        {"trainNum": "L2402", "type": "Local", "depart": "07:20", "arrive": "08:15"},
        {"trainNum": "L2404", "type": "Local", "depart": "09:20", "arrive": "10:15"},
        {"trainNum": "L2406", "type": "Local", "depart": "11:20", "arrive": "12:15"},
        {"trainNum": "L2408", "type": "Local", "depart": "13:20", "arrive": "14:15"},
        {"trainNum": "L2410", "type": "Local", "depart": "15:20", "arrive": "16:15"},
        {"trainNum": "L2412", "type": "Local", "depart": "17:20", "arrive": "18:15"},
    ],
    # dongdaegu ↔ samnangjin (~25 min — local, samnangjin not in KTX timetable)
    "dongdaegu_to_samnangjin": [
        {"trainNum": "L2501", "type": "Local", "depart": "07:10", "arrive": "07:35"},
        {"trainNum": "L2503", "type": "Local", "depart": "08:40", "arrive": "09:05"},
        {"trainNum": "L2505", "type": "Local", "depart": "10:10", "arrive": "10:35"},
        {"trainNum": "L2507", "type": "Local", "depart": "11:40", "arrive": "12:05"},
        {"trainNum": "L2509", "type": "Local", "depart": "13:10", "arrive": "13:35"},
        {"trainNum": "L2511", "type": "Local", "depart": "14:40", "arrive": "15:05"},
        {"trainNum": "L2513", "type": "Local", "depart": "16:10", "arrive": "16:35"},
        {"trainNum": "L2515", "type": "Local", "depart": "17:40", "arrive": "18:05"},
    ],
    "samnangjin_to_dongdaegu": [
        {"trainNum": "L2502", "type": "Local", "depart": "07:40", "arrive": "08:05"},
        {"trainNum": "L2504", "type": "Local", "depart": "09:10", "arrive": "09:35"},
        {"trainNum": "L2506", "type": "Local", "depart": "10:40", "arrive": "11:05"},
        {"trainNum": "L2508", "type": "Local", "depart": "12:10", "arrive": "12:35"},
        {"trainNum": "L2510", "type": "Local", "depart": "13:40", "arrive": "14:05"},
        {"trainNum": "L2512", "type": "Local", "depart": "15:10", "arrive": "15:35"},
        {"trainNum": "L2514", "type": "Local", "depart": "16:40", "arrive": "17:05"},
        {"trainNum": "L2516", "type": "Local", "depart": "18:10", "arrive": "18:35"},
    ],
    # samnangjin ↔ bujeon (~30 min — 경전선 local, KTX goes Miryang→Masan not Bujeon)
    "samnangjin_to_bujeon": [
        {"trainNum": "L2601", "type": "Local", "depart": "07:05", "arrive": "07:35"},
        {"trainNum": "L2603", "type": "Local", "depart": "08:35", "arrive": "09:05"},
        {"trainNum": "L2605", "type": "Local", "depart": "10:05", "arrive": "10:35"},
        {"trainNum": "L2607", "type": "Local", "depart": "11:35", "arrive": "12:05"},
        {"trainNum": "L2609", "type": "Local", "depart": "13:05", "arrive": "13:35"},
        {"trainNum": "L2611", "type": "Local", "depart": "14:35", "arrive": "15:05"},
        {"trainNum": "L2613", "type": "Local", "depart": "16:05", "arrive": "16:35"},
        {"trainNum": "L2615", "type": "Local", "depart": "17:35", "arrive": "18:05"},
    ],
    "bujeon_to_samnangjin": [
        {"trainNum": "L2602", "type": "Local", "depart": "07:40", "arrive": "08:10"},
        {"trainNum": "L2604", "type": "Local", "depart": "09:10", "arrive": "09:40"},
        {"trainNum": "L2606", "type": "Local", "depart": "10:40", "arrive": "11:10"},
        {"trainNum": "L2608", "type": "Local", "depart": "12:10", "arrive": "12:40"},
        {"trainNum": "L2610", "type": "Local", "depart": "13:40", "arrive": "14:10"},
        {"trainNum": "L2612", "type": "Local", "depart": "15:10", "arrive": "15:40"},
        {"trainNum": "L2614", "type": "Local", "depart": "16:40", "arrive": "17:10"},
        {"trainNum": "L2616", "type": "Local", "depart": "18:10", "arrive": "18:40"},
    ],
    # yeongcheon ↔ dongdaegu (~30 min — 중앙선 goes Yeongcheon→Gyeongju, not Dongdaegu)
    "yeongcheon_to_dongdaegu": [
        {"trainNum": "L2701", "type": "Local", "depart": "07:10", "arrive": "07:40"},
        {"trainNum": "L2703", "type": "Local", "depart": "09:10", "arrive": "09:40"},
        {"trainNum": "L2705", "type": "Local", "depart": "11:10", "arrive": "11:40"},
        {"trainNum": "L2707", "type": "Local", "depart": "13:10", "arrive": "13:40"},
        {"trainNum": "L2709", "type": "Local", "depart": "15:10", "arrive": "15:40"},
        {"trainNum": "L2711", "type": "Local", "depart": "17:10", "arrive": "17:40"},
    ],
    "dongdaegu_to_yeongcheon": [
        {"trainNum": "L2702", "type": "Local", "depart": "07:45", "arrive": "08:15"},
        {"trainNum": "L2704", "type": "Local", "depart": "09:45", "arrive": "10:15"},
        {"trainNum": "L2706", "type": "Local", "depart": "11:45", "arrive": "12:15"},
        {"trainNum": "L2708", "type": "Local", "depart": "13:45", "arrive": "14:15"},
        {"trainNum": "L2710", "type": "Local", "depart": "15:45", "arrive": "16:15"},
        {"trainNum": "L2712", "type": "Local", "depart": "17:45", "arrive": "18:15"},
    ],
    # chungju ↔ jecheon (~40 min, local — 중부내륙선 terminates at Chungju; no KTX to Jecheon)
    "chungju_to_jecheon": [
        {"trainNum": "L2801", "type": "Local", "depart": "07:20", "arrive": "08:00"},
        {"trainNum": "L2803", "type": "Local", "depart": "09:20", "arrive": "10:00"},
        {"trainNum": "L2805", "type": "Local", "depart": "11:20", "arrive": "12:00"},
        {"trainNum": "L2807", "type": "Local", "depart": "13:20", "arrive": "14:00"},
        {"trainNum": "L2809", "type": "Local", "depart": "15:20", "arrive": "16:00"},
        {"trainNum": "L2811", "type": "Local", "depart": "17:20", "arrive": "18:00"},
    ],
    "jecheon_to_chungju": [
        {"trainNum": "L2802", "type": "Local", "depart": "07:05", "arrive": "07:45"},
        {"trainNum": "L2804", "type": "Local", "depart": "09:05", "arrive": "09:45"},
        {"trainNum": "L2806", "type": "Local", "depart": "11:05", "arrive": "11:45"},
        {"trainNum": "L2808", "type": "Local", "depart": "13:05", "arrive": "13:45"},
        {"trainNum": "L2810", "type": "Local", "depart": "15:05", "arrive": "15:45"},
        {"trainNum": "L2812", "type": "Local", "depart": "17:05", "arrive": "17:45"},
    ],
    # gwangju_songjeong ↔ suncheon (~50 min, local — different rail lines, no direct KTX)
    "gwangju_songjeong_to_suncheon": [
        {"trainNum": "L2901", "type": "Local", "depart": "07:10", "arrive": "08:00"},
        {"trainNum": "L2903", "type": "Local", "depart": "09:10", "arrive": "10:00"},
        {"trainNum": "L2905", "type": "Local", "depart": "11:10", "arrive": "12:00"},
        {"trainNum": "L2907", "type": "Local", "depart": "13:10", "arrive": "14:00"},
        {"trainNum": "L2909", "type": "Local", "depart": "15:10", "arrive": "16:00"},
        {"trainNum": "L2911", "type": "Local", "depart": "17:10", "arrive": "18:00"},
    ],
    "suncheon_to_gwangju_songjeong": [
        {"trainNum": "L2902", "type": "Local", "depart": "07:15", "arrive": "08:05"},
        {"trainNum": "L2904", "type": "Local", "depart": "09:15", "arrive": "10:05"},
        {"trainNum": "L2906", "type": "Local", "depart": "11:15", "arrive": "12:05"},
        {"trainNum": "L2908", "type": "Local", "depart": "13:15", "arrive": "14:05"},
        {"trainNum": "L2910", "type": "Local", "depart": "15:15", "arrive": "16:05"},
        {"trainNum": "L2912", "type": "Local", "depart": "17:15", "arrive": "18:05"},
    ],
    # yeongcheon ↔ ahwa (~30 min, local)
    "yeongcheon_to_ahwa": [
        {"trainNum": "L2101", "type": "Local", "depart": "07:20", "arrive": "07:50"},
        {"trainNum": "L2103", "type": "Local", "depart": "09:20", "arrive": "09:50"},
        {"trainNum": "L2105", "type": "Local", "depart": "11:20", "arrive": "11:50"},
        {"trainNum": "L2107", "type": "Local", "depart": "13:20", "arrive": "13:50"},
        {"trainNum": "L2109", "type": "Local", "depart": "15:20", "arrive": "15:50"},
        {"trainNum": "L2111", "type": "Local", "depart": "17:20", "arrive": "17:50"},
    ],
    "ahwa_to_yeongcheon": [
        {"trainNum": "L2102", "type": "Local", "depart": "07:55", "arrive": "08:25"},
        {"trainNum": "L2104", "type": "Local", "depart": "09:55", "arrive": "10:25"},
        {"trainNum": "L2106", "type": "Local", "depart": "11:55", "arrive": "12:25"},
        {"trainNum": "L2108", "type": "Local", "depart": "13:55", "arrive": "14:25"},
        {"trainNum": "L2110", "type": "Local", "depart": "15:55", "arrive": "16:25"},
        {"trainNum": "L2112", "type": "Local", "depart": "17:55", "arrive": "18:25"},
    ],
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_time(val):
    """Convert cell value to minutes since midnight. Returns None if no stop."""
    if val is None:
        return None
    if hasattr(val, 'hour'):  # openpyxl datetime.time object
        m = val.hour * 60 + val.minute
        return None if m == 0 else m
    if isinstance(val, str):
        val = val.strip()
        if not val or val.startswith('00:00'):
            return None
        parts = val.split(':')
        if len(parts) >= 2:
            try:
                h, m = int(parts[0]), int(parts[1])
                return None if (h == 0 and m == 0) else h * 60 + m
            except ValueError:
                return None
    return None


def minutes_to_hhmm(m):
    return f"{m // 60:02d}:{m % 60:02d}"


def is_daily(bigo_val):
    if bigo_val is None:
        return True
    s = str(bigo_val).strip()
    return s == '' or '매일' in s


def parse_sheet(ws):
    """
    Parse one sheet and return a flat list of train objects:
      { trainNum, type, stops: [(stationId, minutes_since_midnight), ...] }
    Only includes daily trains. Stops only include game-relevant stations.
    """
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Find header row (contains '열차번호')
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if any(c == '열차번호' for c in row if c is not None):
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    # English name row is always 2 rows after header (Chinese is +1, English is +2)
    eng_row_idx = header_row_idx + 2
    if eng_row_idx >= len(all_rows):
        return []
    english_row = all_rows[eng_row_idx]

    # Build col_idx → station_id mapping from English row
    col_to_station = {}
    for col_idx, val in enumerate(english_row):
        if val is None:
            continue
        if isinstance(val, str):
            name = val.strip()
            if name in STATION_MAPPING:
                col_to_station[col_idx] = STATION_MAPPING[name]

    if not col_to_station:
        return []

    station_cols = sorted(col_to_station.keys())

    # Find 비고 columns from the raw header row — use them as the reliable
    # forward/reverse split boundary (gap heuristic fails when unmapped intermediate
    # stations create a larger in-direction gap than the actual separator).
    raw_header = all_rows[header_row_idx]
    bigo_cols = [i for i, v in enumerate(raw_header) if v == '비고']
    fwd_bigo = bigo_cols[0] if bigo_cols else None
    rev_bigo = bigo_cols[1] if len(bigo_cols) > 1 else None

    if fwd_bigo is not None:
        # Split on first 비고 position
        fwd_station_cols = [(c, col_to_station[c]) for c in station_cols if c < fwd_bigo]
        rev_station_cols = [(c, col_to_station[c]) for c in station_cols if c > fwd_bigo]
        # If a second 비고 exists, further trim rev to only cols before it
        if rev_bigo is not None:
            rev_station_cols = [(c, sid) for c, sid in rev_station_cols if c < rev_bigo]
    else:
        # Fallback: gap heuristic
        best_gap_start = None
        best_gap_len = 0
        for i in range(len(station_cols) - 1):
            gap = station_cols[i + 1] - station_cols[i]
            if gap > best_gap_len:
                best_gap_len = gap
                best_gap_start = station_cols[i]
        if best_gap_start is None or best_gap_len <= 1:
            fwd_station_cols = [(c, col_to_station[c]) for c in station_cols]
            rev_station_cols = []
        else:
            fwd_station_cols = [(c, col_to_station[c]) for c in station_cols if c <= best_gap_start]
            rev_station_cols = [(c, col_to_station[c]) for c in station_cols if c > best_gap_start]

    trains = []
    data_start = header_row_idx + 3  # skip header, Chinese, English rows

    for row in all_rows[data_start:]:
        if row is None:
            continue
        row = list(row)

        # Pad row if shorter than expected
        max_col = max(
            ([c for c, _ in fwd_station_cols] + [c for c, _ in rev_station_cols] + [fwd_bigo or 0, rev_bigo or 0]),
            default=0
        )
        while len(row) <= max_col:
            row.append(None)

        # --- Parse forward train ---
        if fwd_station_cols:
            fwd_first_col = fwd_station_cols[0][0]
            # trainNum is somewhere before the first station col (up to 10 cols back)
            train_num, train_type = None, None
            for offset in range(1, 11):
                c = fwd_first_col - offset
                if c < 0:
                    break
                if isinstance(row[c], (int, float)) and not isinstance(row[c], bool):
                    train_num = int(row[c])
                    # type is one col after trainNum
                    type_col = c + 1
                    if type_col < fwd_first_col and isinstance(row[type_col], str):
                        train_type = row[type_col].strip().replace('_', '-')
                    break

            if train_num is not None and is_daily(row[fwd_bigo] if fwd_bigo else None):
                stops = []
                for col_idx, station_id in fwd_station_cols:
                    t = parse_time(row[col_idx])
                    if t is not None:
                        stops.append((station_id, t))
                if stops:
                    trains.append({
                        'trainNum': train_num,
                        'type': train_type or 'KTX',
                        'stops': stops,
                    })

        # --- Parse reverse train ---
        if rev_station_cols:
            rev_first_col = rev_station_cols[0][0]
            train_num, train_type = None, None
            for offset in range(1, 11):
                c = rev_first_col - offset
                if c < 0:
                    break
                if isinstance(row[c], (int, float)) and not isinstance(row[c], bool):
                    train_num = int(row[c])
                    type_col = c + 1
                    if type_col < rev_first_col and isinstance(row[type_col], str):
                        train_type = row[type_col].strip().replace('_', '-')
                    break

            if train_num is not None and is_daily(row[rev_bigo] if rev_bigo else None):
                stops = []
                for col_idx, station_id in rev_station_cols:
                    t = parse_time(row[col_idx])
                    if t is not None:
                        stops.append((station_id, t))
                if stops:
                    trains.append({
                        'trainNum': train_num,
                        'type': train_type or 'KTX',
                        'stops': stops,
                    })

    return trains


def extract_services(all_trains, from_id, to_id):
    """
    From all parsed trains, extract services that stop at from_id then to_id in order.
    Returns list of { trainNum, type, depart: "HH:MM", arrive: "HH:MM" }
    filtered to departures 07:00–19:00 (420–1140 min).
    """
    services = []
    seen = set()  # deduplicate by (trainNum, depart)

    for train in all_trains:
        stops = train['stops']
        station_ids = [s[0] for s in stops]

        if from_id not in station_ids or to_id not in station_ids:
            continue

        from_idx = station_ids.index(from_id)
        # Find to_id appearing AFTER from_idx
        to_idx = None
        for k in range(from_idx + 1, len(stops)):
            if stops[k][0] == to_id:
                to_idx = k
                break

        if to_idx is None:
            continue

        depart_min = stops[from_idx][1]
        arrive_min = stops[to_idx][1]

        # Filter to game hours: depart 07:00–19:00
        if depart_min < 420 or depart_min > 1140:
            continue

        depart_str = minutes_to_hhmm(depart_min)
        arrive_str = minutes_to_hhmm(arrive_min)
        key = (train['trainNum'], depart_str)
        if key in seen:
            continue
        seen.add(key)

        services.append({
            'trainNum': train['trainNum'],
            'type': train['type'],
            'depart': depart_str,
            'arrive': arrive_str,
        })

    services.sort(key=lambda s: s['depart'])
    return services


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    xlsx_path = os.path.join(os.path.dirname(__file__), '..', 'ktx_timetable.xlsx')
    output_path = os.path.join(os.path.dirname(__file__), '..', 'public', 'data', 'timetable.json')

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading workbook...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("Parsing sheets...")
    all_trains = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        trains = parse_sheet(ws)
        print(f"  {sheet_name}: {len(trains)} trains parsed")
        all_trains.extend(trains)

    print(f"Total trains parsed: {len(all_trains)}")

    # Build timetable for all directed edges
    timetable = {}

    for a, b in SHOW_EDGES:
        key_ab = f"{a}_to_{b}"
        key_ba = f"{b}_to_{a}"

        # Check if we have local services for this edge
        if key_ab in LOCAL_SERVICES:
            timetable[key_ab] = LOCAL_SERVICES[key_ab]
        else:
            svc = extract_services(all_trains, a, b)
            timetable[key_ab] = svc

        if key_ba in LOCAL_SERVICES:
            timetable[key_ba] = LOCAL_SERVICES[key_ba]
        else:
            svc = extract_services(all_trains, b, a)
            timetable[key_ba] = svc

    # Print summary
    print("\nEdge coverage:")
    empty_edges = []
    for key, services in sorted(timetable.items()):
        status = f"{len(services):3d} services"
        print(f"  {key:<40} {status}")
        if len(services) == 0:
            empty_edges.append(key)

    if empty_edges:
        print(f"\nWARNING: {len(empty_edges)} edges have no services:")
        for e in empty_edges:
            print(f"  {e}")

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(timetable, f, ensure_ascii=False, indent=2)

    print(f"\nWrote {output_path}")
    print(f"Total directed edges: {len(timetable)}")
    print(f"Total service entries: {sum(len(v) for v in timetable.values())}")


if __name__ == '__main__':
    main()
